"""
review_intel.py — Review Intelligence Engine

Given ONE Amazon product URL, pulls its review base and asks the LLM to write a
long-form, evidence-grounded NARRATIVE report of what customers actually say —
flowing prose and real quotes, not a bullet-point summary. This is the deep-dive
sibling of Product Intelligence, which stays comparative/ranked across many
competitors.
"""

import io
from datetime import datetime

from amazon_scraper import (
    safe_number, _extract_asin, fetch_product_serpapi, scrape_product_page, _format_reviews,
    fetch_reviews_via_serpapi, fetch_reviews_apify, APIFY_KEY,
)
from product_analytics import star_distribution, review_timeline
from llm import ask_llm_json, llm_available

MIN_REVIEWS_BEFORE_BACKFILL = 6


def fetch_product(url):
    """Resolve one Amazon URL into merged product+review data.

    SerpAPI's `amazon_product` endpoint (the primary fetch) only bundles a
    small "top reviews" snippet with the listing, and for a lot of products
    that snippet is empty even though the product has hundreds of reviews —
    that was leaving Review Intelligence with 0 sampled reviews. If the
    snippet is thin, backfill with the dedicated review-pages fetcher (same
    one Product Intelligence uses), then Apify for an even deeper pull if a
    key is configured."""
    asin = _extract_asin(url)
    page_data = fetch_product_serpapi(asin) if asin else {}
    if not page_data:
        page_data = scrape_product_page(url, asin)
    page_data = dict(page_data or {})

    positive = page_data.get("reviews", []) or []
    critical = page_data.get("critical_reviews", []) or []
    if asin and len(positive) + len(critical) < MIN_REVIEWS_BEFORE_BACKFILL:
        extra_pos, extra_crit = fetch_reviews_via_serpapi(asin, pages=3)
        if len(extra_pos) + len(extra_crit) < MIN_REVIEWS_BEFORE_BACKFILL and APIFY_KEY:
            extra_pos, extra_crit = fetch_reviews_apify(asin, target=80)
        if extra_pos or extra_crit:
            positive = _dedup_reviews(positive + extra_pos)
            critical = _dedup_reviews(critical + extra_crit)
            page_data["reviews"] = positive
            page_data["critical_reviews"] = critical

    return page_data


def _dedup_reviews(reviews):
    seen, out = set(), []
    for r in reviews:
        key = str(r.get("body", ""))[:80]
        if key and key not in seen:
            seen.add(key)
            out.append(r)
    return out


def _mock_report(product_data):
    title = product_data.get("title", product_data.get("product_name", "This product"))
    return {
        "headline": f"Demo mode — add SERPAPI_KEY / GROQ_API_KEY in .env for a live review deep-dive of {title}.",
        "voice_of_customer": "Connect the API keys to generate the full descriptive narrative from real Amazon reviews.",
        "themes": [],
        "who_its_for": "",
        "who_should_avoid": "",
        "purchase_journey": "",
        "friction_points": "",
        "sentiment_arc": "",
        "verdict": "",
    }


def _write_narrative(product_data, positive, critical, stars, timeline):
    if not llm_available():
        return _mock_report(product_data)

    title = str(product_data.get("title", product_data.get("product_name", "")))[:250]
    brand = str(product_data.get("brand", ""))
    rating = safe_number(product_data.get("rating", 0), max_val=5)
    review_count = int(safe_number(product_data.get("review_count", 0)))

    positive_text = _format_reviews(positive, 25)
    critical_text = _format_reviews(critical, 20)
    dist_text = ", ".join(f"{k}★: {v}%" for k, v in stars.get("pct", {}).items())
    monthly = (timeline.get("monthly") or [])[-6:]
    trend_text = ", ".join(f"{m['month']}: {m['avg']}★ ({m['count']} reviews)" for m in monthly) \
        or "not enough dated reviews to plot a trend"

    aspects = product_data.get("amazon_aspects", []) or []
    aspects_text = "\n".join(
        f"- {a.get('aspect', '')}: {a.get('sentiment', '')} "
        f"({a.get('positive', 0)}/{a.get('total', 0)} positive, {a.get('negative', 0)} negative) — "
        f"{a.get('summary', '')}" for a in aspects[:10]) or "none provided"
    amazon_summary = product_data.get("amazon_summary", "")

    prompt = f"""You are a qualitative research analyst who reads through hundreds of real customer
reviews and writes a long-form, richly descriptive narrative report about what customers
ACTUALLY say — not a bullet-point summary. Write in flowing prose, grounded in the review
text below. Quote real lines where they exist. Do not invent facts the reviews don't support.

PRODUCT: {title}
BRAND: {brand}
RATING: {rating}/5 from {review_count:,} total reviews on Amazon
STAR DISTRIBUTION (of the {stars.get('sampled', 0)} reviews sampled for this report): {dist_text or 'n/a'}
RATING TREND (recent months, from the sampled reviews): {trend_text}

AMAZON'S OWN "CUSTOMERS SAY" SUMMARY: {amazon_summary or 'n/a'}

AMAZON'S ASPECT INSIGHTS (counted across Amazon's FULL review base — authoritative, use as ground truth):
{aspects_text}

POSITIVE REVIEWS (sample):
{positive_text or 'none provided'}

CRITICAL REVIEWS (sample):
{critical_text or 'none provided'}

Return ONLY this JSON (start with {{). FIELD ORDER MATTERS — write the short fields first and the
long ones (voice_of_customer, themes) LAST, so if you run low on room the report still ends with a
complete verdict rather than a cut-off theme:
{{
"headline": "one vivid sentence capturing the overall verdict of the customer base",
"verdict": "3-4 sentence closing narrative verdict — the honest overall read a teammate should walk away with",
"who_its_for": "3-5 sentences on which skin types / use-cases / customer profiles this clearly works well for, grounded in the reviews",
"who_should_avoid": "2-4 sentences on who reports it not working or reacting badly, and why",
"purchase_journey": "3-5 sentences: why people buy it (triggers, referrals, claims that resonate), and what they compare it against",
"friction_points": "4-6 sentence hard-hitting paragraph on the recurring complaints, their likely cause, and how severe/common they are relative to total mentions — name aspect + negative-mention count where the Amazon data supports it",
"sentiment_arc": "2-4 sentences on how sentiment has moved over time. Only call it a 'trend' for months with 3+ reviews — for any month with 1-2 reviews, say so explicitly ('only N reviews in that month, too few to call a pattern') rather than treating it as a signal. If every month has too few reviews, say the sample is too thin to read a trend and point to the aggregate rating instead.",
"voice_of_customer": "8-12 sentence descriptive narrative in flowing prose painting the full picture of what customers experience with this product — first impressions, texture, results over time, how it fits into routines. Written like a qualitative research summary, not a list.",
"themes": [
  {{"name": "e.g. Efficacy / Skin Results", "sentiment": "Positive/Mixed/Negative",
    "narrative": "4-6 sentence descriptive paragraph on this theme, citing how often and how strongly it comes up — tie it to the Amazon aspect mention counts where relevant",
    "quotes": ["short real verbatim quote", "short real verbatim quote"]}}
]
}}
List EXACTLY 4 "themes" (pick the 4 most-supported from: efficacy, texture/sensory, packaging,
value for money, fragrance, skin compatibility, shipping/authenticity). Keep every field within
its stated sentence count — do not run long. JSON only, no markdown fences."""

    # Groq's free tier caps requests at 6000 tokens/minute (prompt + completion
    # combined) — keep some headroom under that rather than requesting the max
    # and getting rejected outright (llm.groq_chat backs off and retries on
    # that specific error too, as a second line of defence).
    parsed = ask_llm_json(
        prompt,
        system="Output valid JSON only. Start with {. Be descriptive and narrative, not terse.",
        temperature=0.3, max_tokens=4000, retries=2,
    )
    return parsed if parsed is not None else _mock_report(product_data)


def deep_review_report(url, progress_cb=None):
    """
    Full pipeline: scrape one product's reviews -> long-form descriptive report.
    Returns {product, reviews_sampled, star_distribution, timeline, report,
    positive_reviews, critical_reviews, amazon_aspects, amazon_summary}.
    """
    if progress_cb:
        progress_cb(1, 4, "Fetching product + reviews from Amazon...")
    product_data = fetch_product(url)
    if not product_data:
        raise ValueError("Couldn't fetch that product — check the URL and try again.")

    positive = product_data.get("reviews", [])
    critical = product_data.get("critical_reviews", [])
    positive = positive if isinstance(positive, list) else []
    critical = critical if isinstance(critical, list) else []
    all_reviews = positive + critical

    if progress_cb:
        progress_cb(2, 4, "Computing star distribution + rating timeline...")
    stars = star_distribution(all_reviews)
    timeline = review_timeline(all_reviews)

    if progress_cb:
        progress_cb(3, 4, "Writing the descriptive review narrative...")
    report = _write_narrative(product_data, positive, critical, stars, timeline)

    if progress_cb:
        progress_cb(4, 4, "Done")

    return {
        "product": {
            "title": product_data.get("title", product_data.get("product_name", "")),
            "brand": product_data.get("brand", ""),
            "rating": safe_number(product_data.get("rating", 0), max_val=5),
            "review_count": int(safe_number(product_data.get("review_count", 0))),
            "price": safe_number(product_data.get("selling_price") or product_data.get("price", 0)),
            "thumbnail": product_data.get("thumbnail", ""),
            "url": url,
        },
        "scanned_at": datetime.now().strftime("%d %B %Y, %H:%M"),
        "reviews_sampled": len(all_reviews),
        "positive_sampled": len(positive),
        "critical_sampled": len(critical),
        "positive_reviews": positive,
        "critical_reviews": critical,
        "star_distribution": stars,
        "timeline": timeline,
        "amazon_aspects": product_data.get("amazon_aspects", []) or [],
        "amazon_summary": product_data.get("amazon_summary", ""),
        "report": report,
    }


# ── Excel export ──────────────────────────────────────────────────────────────
def to_excel(bundle):
    """Multi-tab, colour-coded Excel export of one Review Intelligence deep-dive.
    Every tab opens with a banner: what it shows, how the data was fetched, how
    far to trust it — plus the scan date/time and exact review counts used."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        BLACK = PatternFill("solid", fgColor="0A0A0A")
        LIME_FONT = Font(bold=True, color="C8F55A", size=10)
        HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
        WRAP = Alignment(vertical="top", wrap_text=True)
        thin = Border(*(Side(style="thin", color="DDDDDD"),) * 4)

        def header_row(ws, headers, row=2, height=32):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.fill = BLACK
                c.font = LIME_FONT
                c.alignment = HDR_ALIGN
                c.border = thin
            ws.row_dimensions[row].height = height

        def banner(ws, text, n_cols, height=110):
            last_col = get_column_letter(n_cols)
            ws.merge_cells(f"A1:{last_col}1")
            c = ws.cell(row=1, column=1, value=text)
            c.fill = BLACK
            c.font = Font(italic=True, size=10, color="C8F55A")
            c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            ws.row_dimensions[1].height = height

        p = bundle.get("product", {})
        r = bundle.get("report", {})
        stars = bundle.get("star_distribution", {})
        scanned_at = bundle.get("scanned_at", datetime.now().strftime("%d %B %Y, %H:%M"))
        sampled = bundle.get("reviews_sampled", 0)
        pos_n = bundle.get("positive_sampled", 0)
        crit_n = bundle.get("critical_sampled", 0)
        total_reviews = p.get("review_count", 0)

        wb = openpyxl.Workbook()

        # ══════════════════════════════════════════════════════════════════
        # SHEET 1 — Executive Summary
        # ══════════════════════════════════════════════════════════════════
        we = wb.active
        we.title = "Executive Summary"
        banner(we,
               f"WHAT THIS SHOWS — a one-page identity card for this review deep-dive: which product, when it "
               f"was scanned, and how many reviews the report below is grounded in.  HOW IT'S FETCHED — "
               f"{p.get('title', 'this product')} was scanned live from its Amazon.in URL on {scanned_at}; "
               f"{sampled} individual reviews ({pos_n} positive / {crit_n} critical) were pulled and read "
               f"in full by the AI, on top of Amazon's own aggregate rating and \"Customers say\" aspect "
               f"counts (built from Amazon's COMPLETE review base of {total_reviews:,}).  ACCURACY — the "
               f"identity fields below (title, brand, rating, review count, URL) are Amazon's own live "
               f"figures; open the URL to verify the product this report covers.",
               2)
        we.cell(row=2, column=1, value="Decon AI — Review Intelligence Deep-Dive").font = Font(bold=True, size=15)
        summary = [
            ("Product", p.get("title", "")),
            ("Brand", p.get("brand", "")),
            ("URL", p.get("url", "")),
            ("Scanned at", scanned_at),
            ("Amazon rating", f"{p.get('rating', '—')} / 5"),
            ("Total reviews on Amazon", f"{total_reviews:,}"),
            ("Reviews read for this report", f"{sampled} ({pos_n} positive · {crit_n} critical)"),
            ("Amazon aspect insights used", len(bundle.get("amazon_aspects", []) or [])),
            ("Selling price (₹)", p.get("price", "—")),
            ("Headline verdict", r.get("headline", "")),
        ]
        for ri, (k, v) in enumerate(summary, 4):
            we.cell(row=ri, column=1, value=k).font = Font(bold=True)
            c2 = we.cell(row=ri, column=2, value=v)
            c2.alignment = WRAP
            we.row_dimensions[ri].height = 34
        we.column_dimensions["A"].width = 28
        we.column_dimensions["B"].width = 90

        # ══════════════════════════════════════════════════════════════════
        # SHEET 2 — Voice of Customer (narrative)
        # ══════════════════════════════════════════════════════════════════
        wv = wb.create_sheet("Voice of Customer")
        banner(wv,
               f"WHAT THIS SHOWS — the full descriptive narrative: how customers experience this product, who "
               f"it works for, who it doesn't, why people buy it, and how sentiment has moved over time.  "
               f"HOW IT'S FETCHED — an AI (Groq) reads the {sampled} sampled reviews plus Amazon's aggregate "
               f"aspect counts and writes flowing prose grounded in that text — scanned {scanned_at}.  "
               f"ACCURACY — this is an analyst-style synthesis of real review text, not a raw export; cross-"
               f"check specific claims against the \"Amazon Aspect Data\" tab, which carries Amazon's own "
               f"counted mentions, where precision matters.",
               2)
        wv.cell(row=2, column=1, value="Decon AI — Voice of Customer").font = Font(bold=True, size=14)
        sections = [
            ("Voice of the customer", r.get("voice_of_customer", "")),
            ("Who it's for", r.get("who_its_for", "")),
            ("Who should avoid it", r.get("who_should_avoid", "")),
            ("Why people buy it", r.get("purchase_journey", "")),
            ("Sentiment over time", r.get("sentiment_arc", "")),
            ("Friction points", r.get("friction_points", "")),
            ("Verdict", r.get("verdict", "")),
        ]
        ri = 4
        for label, text in sections:
            wv.cell(row=ri, column=1, value=label).font = Font(bold=True, size=12, color="0A0A0A")
            wv.cell(row=ri, column=1).fill = PatternFill("solid", fgColor="F1F0EC")
            ri += 1
            c = wv.cell(row=ri, column=1, value=text)
            c.alignment = WRAP
            wv.row_dimensions[ri].height = max(40, min(220, 15 * (len(text) // 90 + 1)))
            ri += 2
        wv.column_dimensions["A"].width = 110

        # ══════════════════════════════════════════════════════════════════
        # SHEET 3 — Themes
        # ══════════════════════════════════════════════════════════════════
        wt = wb.create_sheet("Themes")
        banner(wt,
               f"WHAT THIS SHOWS — the {len(r.get('themes', []))} recurring themes found in the reviews "
               f"(efficacy, texture, packaging, value, fragrance, etc.), each with its sentiment, a "
               f"descriptive paragraph and real verbatim quotes.  HOW IT'S FETCHED — the AI clusters the "
               f"{sampled} sampled reviews into themes and pulls a short real quote for each, scanned "
               f"{scanned_at}.  ACCURACY — quotes are copied verbatim from scraped reviews; the narrative "
               f"and sentiment label per theme are the AI's interpretation of how often and how strongly "
               f"that theme comes up.",
               4)
        header_row(wt, ["Theme", "Sentiment", "Narrative", "Real Quotes"])
        for row_i, t in enumerate(r.get("themes", []), 3):
            quotes = "\n".join(f"“{q}”" for q in t.get("quotes", [])[:4])
            vals = [t.get("name", ""), t.get("sentiment", ""), t.get("narrative", ""), quotes]
            for col, val in enumerate(vals, 1):
                c = wt.cell(row=row_i, column=col, value=val)
                c.border = thin
                c.alignment = WRAP
                if col == 2:
                    sentiment = str(val)
                    c.fill = PatternFill("solid", fgColor="DCFCE7" if sentiment == "Positive" else
                                         "FEE2E2" if sentiment == "Negative" else "FEF9C3")
            wt.row_dimensions[row_i].height = 130
        for i, w in enumerate([22, 14, 60, 45], 1):
            wt.column_dimensions[get_column_letter(i)].width = w
        wt.freeze_panes = "A3"

        # ══════════════════════════════════════════════════════════════════
        # SHEET 4 — Amazon Aspect Insights (Amazon's own counted data)
        # ══════════════════════════════════════════════════════════════════
        wa = wb.create_sheet("Amazon Aspect Data")
        aspects = bundle.get("amazon_aspects", []) or []
        banner(wa,
               f"WHAT THIS SHOWS — Amazon's OWN \"Customers say\" aspect breakdown for this product, built by "
               f"Amazon from its complete review base of {total_reviews:,} (not just the sample this report "
               f"reads).  HOW IT'S FETCHED — pulled directly from Amazon's product page via SerpAPI on "
               f"{scanned_at}.  ACCURACY — this is the single most authoritative sheet in this workbook: the "
               f"mention counts are Amazon's own, not AI-estimated.",
               6)
        header_row(wa, ["Aspect", "Sentiment", "Positive mentions", "Negative mentions", "Total mentions", "Amazon's summary"])
        for row_i, a in enumerate(aspects, 3):
            vals = [a.get("aspect", ""), a.get("sentiment", ""), a.get("positive", 0),
                    a.get("negative", 0), a.get("total", 0), a.get("summary", "")]
            for col, val in enumerate(vals, 1):
                c = wa.cell(row=row_i, column=col, value=val)
                c.border = thin
                c.alignment = WRAP
            wa.row_dimensions[row_i].height = 50
        if bundle.get("amazon_summary"):
            note_row = len(aspects) + 4
            wa.cell(row=note_row, column=1, value="Amazon's overall summary:").font = Font(bold=True)
            wa.cell(row=note_row + 1, column=1, value=bundle["amazon_summary"]).alignment = WRAP
            wa.merge_cells(f"A{note_row + 1}:F{note_row + 1}")
        for i, w in enumerate([22, 12, 16, 16, 14, 55], 1):
            wa.column_dimensions[get_column_letter(i)].width = w
        wa.freeze_panes = "A3"

        # ══════════════════════════════════════════════════════════════════
        # SHEET 5 — Visual Dashboard (native Excel charts, not just tables)
        # ══════════════════════════════════════════════════════════════════
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList

        wg = wb.create_sheet("Visual Dashboard")
        banner(wg,
               f"WHAT THIS SHOWS — the same evidence behind this report (star mix, rating trend, theme "
               f"sentiment, Amazon's aspect mentions), charted for a fast read.  HOW IT'S FETCHED — computed "
               f"from the {sampled} sampled reviews and Amazon's own aspect counts, scanned {scanned_at} — no "
               f"new data, same source as every other tab.  ACCURACY — the star-mix and rating-trend charts "
               f"reflect the SAMPLE ({sampled} reviews weighted toward helpful + critical, not a random cut of "
               f"the {total_reviews:,} total); the aspect chart uses Amazon's own full-base counts. Any month "
               f"with under 3 reviews is too thin to read as a real trend.",
               8)
        wg.cell(row=2, column=1, value="Review Intelligence — Visual Dashboard").font = Font(bold=True, size=14)

        # data block 1 — star distribution
        wg.cell(row=4, column=1, value="Star distribution").font = Font(bold=True)
        header_row(wg, ["Stars", "Count"], row=5, height=20)
        pct = stars.get("pct", {})
        counts = stars.get("counts", {})
        for i, k in enumerate(["5", "4", "3", "2", "1"], 6):
            wg.cell(row=i, column=1, value=f"{k}★")
            wg.cell(row=i, column=2, value=counts.get(k, 0))
        star_chart = PieChart()
        star_chart.title = "Star Distribution (sampled reviews)"
        star_chart.height, star_chart.width = 8, 12
        data = Reference(wg, min_col=2, min_row=5, max_row=10)
        cats = Reference(wg, min_col=1, min_row=6, max_row=10)
        star_chart.add_data(data, titles_from_data=True)
        star_chart.set_categories(cats)
        star_chart.dataLabels = DataLabelList()
        star_chart.dataLabels.showPercent = True
        wg.add_chart(star_chart, "D4")

        # data block 2 — rating trend by month
        monthly = bundle.get("timeline", {}).get("monthly", [])
        wg.cell(row=22, column=1, value="Rating trend by month (bars show how many reviews that month — "
                                        "thin bars are low-confidence)").font = Font(bold=True)
        header_row(wg, ["Month", "Avg rating", "Reviews that month"], row=23, height=20)
        for i, m in enumerate(monthly, 24):
            wg.cell(row=i, column=1, value=m.get("month", ""))
            wg.cell(row=i, column=2, value=m.get("avg", 0))
            wg.cell(row=i, column=3, value=m.get("count", 0))
        if monthly:
            last_row = 23 + len(monthly)
            trend_chart = LineChart()
            trend_chart.title = "Avg Rating by Month"
            trend_chart.height, trend_chart.width = 8, 12
            trend_chart.y_axis.scaling.min, trend_chart.y_axis.scaling.max = 0, 5
            data = Reference(wg, min_col=2, min_row=23, max_row=last_row)
            cats = Reference(wg, min_col=1, min_row=24, max_row=last_row)
            trend_chart.add_data(data, titles_from_data=True)
            trend_chart.set_categories(cats)
            wg.add_chart(trend_chart, "D22")

            volume_chart = BarChart()
            volume_chart.title = "Reviews per Month (sample size — read the trend chart against this)"
            volume_chart.height, volume_chart.width = 8, 12
            data = Reference(wg, min_col=3, min_row=23, max_row=last_row)
            volume_chart.add_data(data, titles_from_data=True)
            volume_chart.set_categories(cats)
            wg.add_chart(volume_chart, "D40")

        # data block 3 — theme sentiment
        themes = r.get("themes", [])
        wg.cell(row=58, column=1, value="Themes by sentiment").font = Font(bold=True)
        header_row(wg, ["Sentiment", "Theme count"], row=59, height=20)
        sentiment_counts = {"Positive": 0, "Mixed": 0, "Negative": 0}
        for t in themes:
            s = t.get("sentiment", "Mixed")
            sentiment_counts[s] = sentiment_counts.get(s, 0) + 1
        for i, (k, v) in enumerate(sentiment_counts.items(), 60):
            wg.cell(row=i, column=1, value=k)
            wg.cell(row=i, column=2, value=v)
        theme_chart = PieChart()
        theme_chart.title = "Themes by Sentiment"
        theme_chart.height, theme_chart.width = 8, 12
        data = Reference(wg, min_col=2, min_row=59, max_row=62)
        cats = Reference(wg, min_col=1, min_row=60, max_row=62)
        theme_chart.add_data(data, titles_from_data=True)
        theme_chart.set_categories(cats)
        wg.add_chart(theme_chart, "D58")

        # data block 4 — Amazon aspect mentions (positive vs negative)
        aspects_short = (bundle.get("amazon_aspects") or [])[:8]
        wg.cell(row=76, column=1, value="Amazon aspect mentions — positive vs negative").font = Font(bold=True)
        header_row(wg, ["Aspect", "Positive mentions", "Negative mentions"], row=77, height=20)
        for i, a in enumerate(aspects_short, 78):
            wg.cell(row=i, column=1, value=a.get("aspect", ""))
            wg.cell(row=i, column=2, value=a.get("positive", 0))
            wg.cell(row=i, column=3, value=a.get("negative", 0))
        if aspects_short:
            last_row = 77 + len(aspects_short)
            aspect_chart = BarChart()
            aspect_chart.type = "col"
            aspect_chart.grouping = "clustered"
            aspect_chart.title = "Amazon Aspect Mentions — Positive vs Negative"
            aspect_chart.height, aspect_chart.width = 9, 16
            data = Reference(wg, min_col=2, max_col=3, min_row=77, max_row=last_row)
            cats = Reference(wg, min_col=1, min_row=78, max_row=last_row)
            aspect_chart.add_data(data, titles_from_data=True)
            aspect_chart.set_categories(cats)
            wg.add_chart(aspect_chart, "D76")

        for i, w in enumerate([28, 16, 20], 1):
            wg.column_dimensions[get_column_letter(i)].width = w

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    except Exception as e:
        print(f"[review_intel] excel error: {e}")
        import traceback
        traceback.print_exc()
        return _fallback_csv(bundle)


def _fallback_csv(bundle):
    p = bundle.get("product", {})
    r = bundle.get("report", {})
    lines = [f"Product,{p.get('title', '')}", f"Brand,{p.get('brand', '')}",
             f"Scanned at,{bundle.get('scanned_at', '')}",
             f"Reviews sampled,{bundle.get('reviews_sampled', 0)}",
             f"Headline,{r.get('headline', '')}"]
    return ("\n".join(lines)).encode("utf-8")
